"""Utility functions for budget import/export"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from .models import Budget, BudgetEntry
from decimal import Decimal


def export_budget_to_excel(budget):
    """Export a budget to Excel format matching the original Numbers structure"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{budget.name} {budget.year}"

    # Month headers
    months = ['Januar', 'Februar', 'März', 'April', 'Mai', 'Juni',
              'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']

    # Header row
    ws['A1'] = 'Kategorie'
    ws['A1'].font = Font(bold=True)
    ws['A1'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for idx, month in enumerate(months, start=2):
        cell = ws.cell(row=1, column=idx)
        cell.value = month
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Total column
    total_col = len(months) + 2
    ws.cell(row=1, column=total_col).value = 'Gesamt'
    ws.cell(row=1, column=total_col).font = Font(bold=True)
    ws.cell(row=1, column=total_col).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Get all categories and entries
    categories = budget.categories.filter(is_active=True).order_by('order')
    entries = BudgetEntry.objects.filter(category__budget=budget, year=budget.year)

    current_row = 2

    # Group categories by type
    category_types = {
        'INCOME': 'Einnahmen',
        'FIXED_EXPENSE': 'Fixkosten',
        'VARIABLE_EXPENSE': 'Variable Kosten',
        'SAVINGS': 'Sparen'
    }

    for cat_type, type_label in category_types.items():
        type_categories = [c for c in categories if c.category_type == cat_type]
        if not type_categories:
            continue

        # Type header
        ws.cell(row=current_row, column=1).value = type_label
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        current_row += 1

        for category in type_categories:
            ws.cell(row=current_row, column=1).value = category.name

            # Get entries for this category
            category_entries = {e.month: e for e in entries if e.category_id == category.id}
            yearly_total = Decimal('0')

            for month in range(1, 13):
                entry = category_entries.get(month)
                cell = ws.cell(row=current_row, column=month + 1)

                if entry:
                    amount = Decimal(entry.actual_amount or entry.planned_amount)
                    cell.value = float(amount)
                    cell.number_format = '#,##0.00'
                    yearly_total += amount

                    # Color coding based on status
                    if entry.actual_amount:
                        if entry.status == 'WITHIN_BUDGET':
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif entry.status == 'WARNING':
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        elif entry.status == 'OVER_BUDGET':
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    cell.value = 0
                    cell.number_format = '#,##0.00'

                cell.alignment = Alignment(horizontal='right')

            # Total column
            total_cell = ws.cell(row=current_row, column=total_col)
            total_cell.value = float(yearly_total)
            total_cell.number_format = '#,##0.00'
            total_cell.font = Font(bold=True)
            total_cell.alignment = Alignment(horizontal='right')

            current_row += 1

        current_row += 1  # Space between category groups

    # Summary row
    ws.cell(row=current_row, column=1).value = 'BILANZ'
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)

    # Calculate monthly and yearly totals
    for month in range(1, 13):
        month_income = sum(
            Decimal(e.actual_amount or e.planned_amount)
            for e in entries
            if e.month == month and e.category.category_type == 'INCOME'
        )
        month_expenses = sum(
            Decimal(e.actual_amount or e.planned_amount)
            for e in entries
            if e.month == month and e.category.category_type != 'INCOME'
        )
        balance = month_income - month_expenses

        cell = ws.cell(row=current_row, column=month + 1)
        cell.value = float(balance)
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')

        if balance >= 0:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Yearly balance
    yearly_summary = budget.get_yearly_summary()
    balance_cell = ws.cell(row=current_row, column=total_col)
    balance_cell.value = float(yearly_summary['balance'])
    balance_cell.number_format = '#,##0.00'
    balance_cell.font = Font(bold=True, size=12)
    balance_cell.alignment = Alignment(horizontal='right')

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 25
    for col in range(2, total_col + 1):
        ws.column_dimensions[chr(64 + col)].width = 12

    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{budget.name}_{budget.year}.xlsx"'

    wb.save(response)
    return response
